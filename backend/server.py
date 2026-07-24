from fastapi import FastAPI, APIRouter, HTTPException, Request, Depends, Query
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import io
import secrets
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Literal
import uuid
from datetime import datetime, timezone, timedelta
import httpx
from openpyxl import Workbook


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]

app = FastAPI()
api_router = APIRouter(prefix="/api")

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)


# =================== AUTH MODELS ===================

class User(BaseModel):
    user_id: str
    email: str
    name: str = ""
    picture: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserPublic(BaseModel):
    user_id: str
    email: str
    name: str
    picture: str


# =================== AUTH DEPENDENCY ===================

async def get_current_user(request: Request) -> User:
    auth = request.headers.get("Authorization") or ""
    if not auth.startswith("Bearer "):
        raise HTTPException(401, "missing bearer token")
    token = auth[7:].strip()
    sess = await db.user_sessions.find_one({"session_token": token}, {"_id": 0})
    if not sess:
        raise HTTPException(401, "invalid session")
    exp = sess.get("expires_at")
    if exp:
        if isinstance(exp, str):
            exp = datetime.fromisoformat(exp)
        if exp.tzinfo is None:
            exp = exp.replace(tzinfo=timezone.utc)
        if exp < datetime.now(timezone.utc):
            raise HTTPException(401, "session expired")
    user_doc = await db.users.find_one({"user_id": sess["user_id"]}, {"_id": 0})
    if not user_doc:
        raise HTTPException(401, "user not found")
    return User(**user_doc)


# =================== AUTH ROUTES ===================

class SessionRequest(BaseModel):
    session_token: str

@api_router.post("/auth/session")
async def create_session(body: SessionRequest):
    """Exchange Emergent session_token for local user; store in user_sessions."""
    async with httpx.AsyncClient(timeout=15) as http:
        r = await http.get(
            "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
            headers={"X-Session-ID": body.session_token},
        )
    if r.status_code != 200:
        raise HTTPException(401, f"invalid session_token: {r.text}")
    data = r.json()
    email = data["email"]
    name = data.get("name", "")
    picture = data.get("picture", "")

    # Upsert user by email
    existing = await db.users.find_one({"email": email}, {"_id": 0})
    if existing:
        user_id = existing["user_id"]
        await db.users.update_one({"email": email}, {"$set": {"name": name, "picture": picture}})
    else:
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        await db.users.insert_one(User(user_id=user_id, email=email, name=name, picture=picture).dict())
        # Claim any unclaimed (owner_id=None) legacy data for the first user
        first_user = await db.users.count_documents({}) == 1
        if first_user:
            for coll in ["categories", "transactions", "accounts", "allocations", "budgets", "balance_snapshots"]:
                await db[coll].update_many(
                    {"$or": [{"owner_id": None}, {"owner_id": {"$exists": False}}]},
                    {"$set": {"owner_id": user_id}},
                )

    session_token = data.get("session_token") or secrets.token_urlsafe(32)
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    await db.user_sessions.insert_one({
        "session_token": session_token,
        "user_id": user_id,
        "expires_at": expires_at,
        "created_at": datetime.now(timezone.utc),
    })

    return {
        "session_token": session_token,
        "user": {"user_id": user_id, "email": email, "name": name, "picture": picture},
    }

@api_router.get("/auth/me", response_model=UserPublic)
async def auth_me(user: User = Depends(get_current_user)):
    return UserPublic(**user.dict())

@api_router.post("/auth/logout")
async def logout(request: Request):
    auth = request.headers.get("Authorization") or ""
    if auth.startswith("Bearer "):
        token = auth[7:].strip()
        await db.user_sessions.delete_one({"session_token": token})
    return {"ok": True}


# =================== DATA MODELS ===================

class Category(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    owner_id: str
    name: str
    type: Literal["expense", "income"]
    planned: float = 0.0
    month: str

class CategoryCreate(BaseModel):
    name: str
    type: Literal["expense", "income"]
    planned: float = 0.0
    month: str

class CategoryUpdate(BaseModel):
    name: Optional[str] = None
    planned: Optional[float] = None


class Transaction(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    owner_id: str
    date: str
    amount: float
    description: str = ""
    category: str
    type: Literal["expense", "income", "transfer"]
    month: str
    account_id: Optional[str] = None
    to_account_id: Optional[str] = None  # for transfers

class TransactionCreate(BaseModel):
    date: str
    amount: float
    description: str = ""
    category: str = ""
    type: Literal["expense", "income", "transfer"] = "expense"
    account_id: Optional[str] = None
    to_account_id: Optional[str] = None

class TransactionUpdate(BaseModel):
    date: Optional[str] = None
    amount: Optional[float] = None
    description: Optional[str] = None
    category: Optional[str] = None
    account_id: Optional[str] = None


class Account(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    owner_id: str
    name: str
    group: str
    balance: float
    brought_forward: float = 0.0

class AccountCreate(BaseModel):
    name: str
    group: str
    balance: float = 0.0
    brought_forward: float = 0.0

class AccountUpdate(BaseModel):
    name: Optional[str] = None
    group: Optional[str] = None
    balance: Optional[float] = None


class Allocation(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    owner_id: str
    name: str
    percent: float

class AllocationUpdate(BaseModel):
    percent: float


class BalanceSnapshot(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    owner_id: str
    account_id: str
    month: str  # YYYY-MM (end-of-month balance)
    balance: float
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class BalanceSnapshotCreate(BaseModel):
    account_id: str
    month: str
    balance: float


class Budget(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    owner_id: str
    month: str
    name: str
    share_token: Optional[str] = None
    share_write: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class BudgetCreate(BaseModel):
    month: str
    name: str

class BudgetShare(BaseModel):
    write: bool = False


# =================== HELPERS ===================

def _dict(model) -> dict:
    d = model.dict()
    return d


# =================== CATEGORY ROUTES ===================

@api_router.get("/categories", response_model=List[Category])
async def list_categories(month: Optional[str] = None, user: User = Depends(get_current_user)):
    q: dict = {"owner_id": user.user_id}
    if month:
        q["month"] = month
    docs = await db.categories.find(q, {"_id": 0}).to_list(2000)
    return [Category(**d) for d in docs]

@api_router.post("/categories", response_model=Category)
async def create_category(body: CategoryCreate, user: User = Depends(get_current_user)):
    obj = Category(owner_id=user.user_id, **body.dict())
    await db.categories.insert_one(obj.dict())
    return obj

@api_router.put("/categories/{cat_id}", response_model=Category)
async def update_category(cat_id: str, body: CategoryUpdate, user: User = Depends(get_current_user)):
    update = {k: v for k, v in body.dict().items() if v is not None}
    if not update:
        raise HTTPException(400, "no fields")
    await db.categories.update_one({"id": cat_id, "owner_id": user.user_id}, {"$set": update})
    doc = await db.categories.find_one({"id": cat_id, "owner_id": user.user_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "not found")
    return Category(**doc)

@api_router.delete("/categories/{cat_id}")
async def delete_category(cat_id: str, user: User = Depends(get_current_user)):
    await db.categories.delete_one({"id": cat_id, "owner_id": user.user_id})
    return {"ok": True}


# =================== TRANSACTION ROUTES ===================

@api_router.get("/transactions", response_model=List[Transaction])
async def list_transactions(
    month: Optional[str] = None,
    type: Optional[str] = None,
    account_id: Optional[str] = None,
    user: User = Depends(get_current_user),
):
    q: dict = {"owner_id": user.user_id}
    if month:
        q["month"] = month
    if type:
        q["type"] = type
    if account_id:
        q["$or"] = [{"account_id": account_id}, {"to_account_id": account_id}]
    docs = await db.transactions.find(q, {"_id": 0}).sort("date", -1).to_list(3000)
    return [Transaction(**d) for d in docs]

async def _apply_tx_to_account(tx: Transaction, direction: int = 1):
    """Adjust account balances by tx. direction=1 for apply, -1 for revert."""
    amt = tx.amount * direction
    if tx.type == "expense" and tx.account_id:
        await db.accounts.update_one(
            {"id": tx.account_id, "owner_id": tx.owner_id}, {"$inc": {"balance": -amt}}
        )
    elif tx.type == "income" and tx.account_id:
        await db.accounts.update_one(
            {"id": tx.account_id, "owner_id": tx.owner_id}, {"$inc": {"balance": amt}}
        )
    elif tx.type == "transfer" and tx.account_id and tx.to_account_id:
        await db.accounts.update_one(
            {"id": tx.account_id, "owner_id": tx.owner_id}, {"$inc": {"balance": -amt}}
        )
        await db.accounts.update_one(
            {"id": tx.to_account_id, "owner_id": tx.owner_id}, {"$inc": {"balance": amt}}
        )

@api_router.post("/transactions", response_model=Transaction)
async def create_transaction(body: TransactionCreate, user: User = Depends(get_current_user)):
    month = body.date[:7]
    if body.type == "transfer":
        if not body.account_id or not body.to_account_id:
            raise HTTPException(400, "transfer requires from and to account")
        payload = body.dict()
        payload["category"] = payload.get("category") or "Transfer"
        obj = Transaction(owner_id=user.user_id, month=month, **payload)
    else:
        obj = Transaction(owner_id=user.user_id, month=month, **body.dict())
    await db.transactions.insert_one(obj.dict())
    await _apply_tx_to_account(obj, 1)
    return obj

@api_router.put("/transactions/{tx_id}", response_model=Transaction)
async def update_transaction(tx_id: str, body: TransactionUpdate, user: User = Depends(get_current_user)):
    old = await db.transactions.find_one({"id": tx_id, "owner_id": user.user_id}, {"_id": 0})
    if not old:
        raise HTTPException(404, "not found")
    update = {k: v for k, v in body.dict().items() if v is not None}
    if "date" in update:
        update["month"] = update["date"][:7]
    if not update:
        raise HTTPException(400, "no fields")
    # revert old effect
    await _apply_tx_to_account(Transaction(**old), -1)
    await db.transactions.update_one({"id": tx_id, "owner_id": user.user_id}, {"$set": update})
    doc = await db.transactions.find_one({"id": tx_id, "owner_id": user.user_id}, {"_id": 0})
    # apply new effect
    await _apply_tx_to_account(Transaction(**doc), 1)
    return Transaction(**doc)

@api_router.delete("/transactions/{tx_id}")
async def delete_transaction(tx_id: str, user: User = Depends(get_current_user)):
    old = await db.transactions.find_one({"id": tx_id, "owner_id": user.user_id}, {"_id": 0})
    if old:
        await _apply_tx_to_account(Transaction(**old), -1)
        await db.transactions.delete_one({"id": tx_id, "owner_id": user.user_id})
    return {"ok": True}


# =================== ACCOUNT ROUTES ===================

@api_router.get("/accounts", response_model=List[Account])
async def list_accounts(user: User = Depends(get_current_user)):
    docs = await db.accounts.find({"owner_id": user.user_id}, {"_id": 0}).to_list(1000)
    return [Account(**d) for d in docs]

@api_router.post("/accounts", response_model=Account)
async def create_account(body: AccountCreate, user: User = Depends(get_current_user)):
    obj = Account(owner_id=user.user_id, **body.dict())
    await db.accounts.insert_one(obj.dict())
    return obj

@api_router.put("/accounts/{acc_id}", response_model=Account)
async def update_account(acc_id: str, body: AccountUpdate, user: User = Depends(get_current_user)):
    update = {k: v for k, v in body.dict().items() if v is not None}
    if not update:
        raise HTTPException(400, "no fields")
    await db.accounts.update_one({"id": acc_id, "owner_id": user.user_id}, {"$set": update})
    doc = await db.accounts.find_one({"id": acc_id, "owner_id": user.user_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "not found")
    return Account(**doc)

@api_router.delete("/accounts/{acc_id}")
async def delete_account(acc_id: str, user: User = Depends(get_current_user)):
    await db.accounts.delete_one({"id": acc_id, "owner_id": user.user_id})
    return {"ok": True}


# =================== ALLOCATION ROUTES ===================

@api_router.get("/allocations", response_model=List[Allocation])
async def list_allocations(user: User = Depends(get_current_user)):
    docs = await db.allocations.find({"owner_id": user.user_id}, {"_id": 0}).to_list(100)
    return [Allocation(**d) for d in docs]

@api_router.put("/allocations/{alloc_id}", response_model=Allocation)
async def update_allocation(alloc_id: str, body: AllocationUpdate, user: User = Depends(get_current_user)):
    await db.allocations.update_one(
        {"id": alloc_id, "owner_id": user.user_id}, {"$set": {"percent": body.percent}}
    )
    doc = await db.allocations.find_one({"id": alloc_id, "owner_id": user.user_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "not found")
    return Allocation(**doc)


# =================== BALANCE SNAPSHOTS ===================

@api_router.get("/balance-snapshots", response_model=List[BalanceSnapshot])
async def list_snapshots(month: Optional[str] = None, user: User = Depends(get_current_user)):
    q: dict = {"owner_id": user.user_id}
    if month:
        q["month"] = month
    docs = await db.balance_snapshots.find(q, {"_id": 0}).sort("month", 1).to_list(2000)
    return [BalanceSnapshot(**d) for d in docs]

@api_router.post("/balance-snapshots", response_model=BalanceSnapshot)
async def create_snapshot(body: BalanceSnapshotCreate, user: User = Depends(get_current_user)):
    """Upsert an end-of-month balance for an account.

    Also updates the account's current balance to this reported value and
    stores the previous balance as brought_forward for gain calculation.
    """
    account = await db.accounts.find_one(
        {"id": body.account_id, "owner_id": user.user_id}, {"_id": 0}
    )
    if not account:
        raise HTTPException(404, "account not found")

    existing = await db.balance_snapshots.find_one(
        {"owner_id": user.user_id, "account_id": body.account_id, "month": body.month},
        {"_id": 0},
    )
    if existing:
        await db.balance_snapshots.update_one(
            {"id": existing["id"]}, {"$set": {"balance": body.balance}}
        )
        obj = BalanceSnapshot(**{**existing, "balance": body.balance})
    else:
        obj = BalanceSnapshot(owner_id=user.user_id, **body.dict())
        await db.balance_snapshots.insert_one(obj.dict())

    # update account balance
    await db.accounts.update_one(
        {"id": body.account_id, "owner_id": user.user_id},
        {"$set": {"balance": body.balance}},
    )
    return obj


# =================== ROLLOVER ===================

class RolloverBody(BaseModel):
    from_month: str
    to_month: str

@api_router.post("/rollover")
async def rollover(body: RolloverBody, user: User = Depends(get_current_user)):
    """Carry forward: create categories for to_month copied from from_month, and set
    brought_forward on accounts equal to current balance."""
    # Copy categories
    src_cats = await db.categories.find(
        {"owner_id": user.user_id, "month": body.from_month}, {"_id": 0}
    ).to_list(1000)
    existing = await db.categories.count_documents(
        {"owner_id": user.user_id, "month": body.to_month}
    )
    if existing == 0:
        for c in src_cats:
            new_cat = Category(
                owner_id=user.user_id,
                name=c["name"],
                type=c["type"],
                planned=c["planned"],
                month=body.to_month,
            )
            await db.categories.insert_one(new_cat.dict())

    # Set brought_forward to current balance for all accounts
    accounts = await db.accounts.find({"owner_id": user.user_id}, {"_id": 0}).to_list(1000)
    for a in accounts:
        await db.accounts.update_one(
            {"id": a["id"], "owner_id": user.user_id},
            {"$set": {"brought_forward": a["balance"]}},
        )
    return {"ok": True, "categories_created": len(src_cats) if existing == 0 else 0}


# =================== SUMMARY ===================

async def _compute_summary(user_id: str, month: str) -> dict:
    cats = await db.categories.find({"owner_id": user_id, "month": month}, {"_id": 0}).to_list(1000)
    txs = await db.transactions.find({"owner_id": user_id, "month": month}, {"_id": 0}).to_list(3000)
    accounts = await db.accounts.find({"owner_id": user_id}, {"_id": 0}).to_list(1000)

    non_transfer_txs = [t for t in txs if t["type"] != "transfer"]
    planned_expense = sum(c["planned"] for c in cats if c["type"] == "expense")
    planned_income = sum(c["planned"] for c in cats if c["type"] == "income")
    actual_expense = sum(t["amount"] for t in non_transfer_txs if t["type"] == "expense")
    actual_income = sum(t["amount"] for t in non_transfer_txs if t["type"] == "income")

    starting_balance = sum(a.get("brought_forward", 0) for a in accounts)
    saved_this_month = actual_income - actual_expense
    end_balance = sum(a.get("balance", 0) for a in accounts)

    cat_actual: dict = {}
    for t in non_transfer_txs:
        cat_actual[t["category"]] = cat_actual.get(t["category"], 0) + t["amount"]

    categories_with_actual = []
    for c in cats:
        categories_with_actual.append({
            **c,
            "actual": round(cat_actual.get(c["name"], 0), 2),
            "diff": round(c["planned"] - cat_actual.get(c["name"], 0), 2),
        })

    return {
        "month": month,
        "starting_balance": round(starting_balance, 2),
        "end_balance": round(end_balance, 2),
        "saved_this_month": round(saved_this_month, 2),
        "planned_expense": round(planned_expense, 2),
        "actual_expense": round(actual_expense, 2),
        "planned_income": round(planned_income, 2),
        "actual_income": round(actual_income, 2),
        "categories": categories_with_actual,
    }

@api_router.get("/summary")
async def get_summary(month: str, user: User = Depends(get_current_user)):
    return await _compute_summary(user.user_id, month)

@api_router.get("/months")
async def get_months(user: User = Depends(get_current_user)):
    tx_months = await db.transactions.distinct("month", {"owner_id": user.user_id})
    cat_months = await db.categories.distinct("month", {"owner_id": user.user_id})
    months = sorted(set(tx_months + cat_months), reverse=True)
    if not months:
        months = [datetime.now(timezone.utc).strftime("%Y-%m")]
    return {"months": months}


# =================== CHARTS ===================

@api_router.get("/charts")
async def get_charts(user: User = Depends(get_current_user)):
    """Return net worth over months, planned vs actual per current month, spending by category."""
    accounts = await db.accounts.find({"owner_id": user.user_id}, {"_id": 0}).to_list(1000)
    snaps = await db.balance_snapshots.find({"owner_id": user.user_id}, {"_id": 0}).sort("month", 1).to_list(2000)

    # net worth per month = sum of latest snapshot per account up to that month, else brought_forward
    months = sorted(set(s["month"] for s in snaps))
    net_worth_series = []
    for m in months:
        total = 0.0
        for a in accounts:
            # latest snapshot for a at or before m
            per = [s for s in snaps if s["account_id"] == a["id"] and s["month"] <= m]
            if per:
                total += per[-1]["balance"]
            else:
                total += a.get("brought_forward", 0)
        net_worth_series.append({"month": m, "value": round(total, 2)})

    # current total net worth
    net_worth_series.append({
        "month": "now",
        "value": round(sum(a["balance"] for a in accounts), 2),
    })

    # spending by category for current month
    now_month = datetime.now(timezone.utc).strftime("%Y-%m")
    tx_months = await db.transactions.distinct("month", {"owner_id": user.user_id})
    latest_month = max(tx_months) if tx_months else now_month
    txs = await db.transactions.find(
        {"owner_id": user.user_id, "month": latest_month, "type": "expense"}, {"_id": 0}
    ).to_list(3000)
    by_cat: dict = {}
    for t in txs:
        by_cat[t["category"]] = by_cat.get(t["category"], 0) + t["amount"]
    spending_by_category = [
        {"label": k, "value": round(v, 2)} for k, v in sorted(by_cat.items(), key=lambda x: -x[1])
    ]

    summary = await _compute_summary(user.user_id, latest_month)
    planned_vs_actual = [
        {
            "label": c["name"],
            "planned": c["planned"],
            "actual": c.get("actual", 0),
        }
        for c in summary["categories"] if c["type"] == "expense"
    ]

    return {
        "net_worth_series": net_worth_series,
        "spending_by_category": spending_by_category,
        "planned_vs_actual": planned_vs_actual,
        "latest_month": latest_month,
    }


# =================== BUDGETS ===================

@api_router.get("/budgets", response_model=List[Budget])
async def list_budgets(user: User = Depends(get_current_user)):
    docs = await db.budgets.find({"owner_id": user.user_id}, {"_id": 0}).sort("month", -1).to_list(200)
    return [Budget(**d) for d in docs]

@api_router.post("/budgets", response_model=Budget)
async def create_budget(body: BudgetCreate, user: User = Depends(get_current_user)):
    obj = Budget(owner_id=user.user_id, **body.dict())
    await db.budgets.insert_one(obj.dict())
    return obj

@api_router.delete("/budgets/{bid}")
async def delete_budget(bid: str, user: User = Depends(get_current_user)):
    await db.budgets.delete_one({"id": bid, "owner_id": user.user_id})
    return {"ok": True}

@api_router.post("/budgets/{bid}/share")
async def share_budget(bid: str, body: BudgetShare, user: User = Depends(get_current_user)):
    doc = await db.budgets.find_one({"id": bid, "owner_id": user.user_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "not found")
    token = doc.get("share_token") or secrets.token_urlsafe(16)
    await db.budgets.update_one(
        {"id": bid, "owner_id": user.user_id},
        {"$set": {"share_token": token, "share_write": body.write}},
    )
    return {"share_token": token, "write": body.write}

@api_router.delete("/budgets/{bid}/share")
async def unshare_budget(bid: str, user: User = Depends(get_current_user)):
    await db.budgets.update_one(
        {"id": bid, "owner_id": user.user_id},
        {"$set": {"share_token": None, "share_write": False}},
    )
    return {"ok": True}


@api_router.get("/shared/{token}")
async def get_shared_budget(token: str):
    b = await db.budgets.find_one({"share_token": token}, {"_id": 0})
    if not b:
        raise HTTPException(404, "not found")
    summary = await _compute_summary(b["owner_id"], b["month"])
    return {
        "budget": b,
        "summary": summary,
        "write": b.get("share_write", False),
    }


class SharedUpdate(BaseModel):
    category_id: str
    planned: float

@api_router.put("/shared/{token}/category")
async def update_shared_category(token: str, body: SharedUpdate):
    b = await db.budgets.find_one({"share_token": token}, {"_id": 0})
    if not b:
        raise HTTPException(404, "not found")
    if not b.get("share_write"):
        raise HTTPException(403, "read-only share")
    await db.categories.update_one(
        {"id": body.category_id, "owner_id": b["owner_id"], "month": b["month"]},
        {"$set": {"planned": body.planned}},
    )
    return {"ok": True}


# =================== EXCEL EXPORT ===================

@api_router.get("/export/excel")
async def export_excel(month: str, user: User = Depends(get_current_user)):
    """Export a month's budget (Summary + Transactions + Accounts) as .xlsx"""
    summary = await _compute_summary(user.user_id, month)
    txs = await db.transactions.find(
        {"owner_id": user.user_id, "month": month}, {"_id": 0}
    ).sort("date", 1).to_list(3000)
    accounts = await db.accounts.find({"owner_id": user.user_id}, {"_id": 0}).to_list(1000)

    wb = Workbook()

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    ws.append([f"Budget summary — {month}"])
    ws.append([])
    ws.append(["Starting balance", summary["starting_balance"]])
    ws.append(["End balance", summary["end_balance"]])
    ws.append(["Saved this month", summary["saved_this_month"]])
    ws.append([])
    ws.append(["", "Planned", "Actual", "Diff"])
    ws.append(["Expenses total", summary["planned_expense"], summary["actual_expense"], summary["planned_expense"] - summary["actual_expense"]])
    ws.append(["Income total", summary["planned_income"], summary["actual_income"], summary["planned_income"] - summary["actual_income"]])
    ws.append([])
    ws.append(["Category", "Type", "Planned", "Actual", "Diff"])
    for c in summary["categories"]:
        ws.append([c["name"], c["type"], c["planned"], c.get("actual", 0), c.get("diff", 0)])

    # Transactions sheet
    ws2 = wb.create_sheet("Transactions")
    ws2.append(["Date", "Type", "Category", "Description", "Amount", "Account", "To Account"])
    account_names = {a["id"]: a["name"] for a in accounts}
    for t in txs:
        ws2.append([
            t["date"], t["type"], t.get("category", ""), t.get("description", ""),
            t["amount"],
            account_names.get(t.get("account_id"), ""),
            account_names.get(t.get("to_account_id"), ""),
        ])

    # Accounts sheet
    ws3 = wb.create_sheet("Accounts")
    ws3.append(["Name", "Group", "Balance", "Brought forward", "Change"])
    for a in accounts:
        ws3.append([a["name"], a["group"], a["balance"], a.get("brought_forward", 0),
                    a["balance"] - a.get("brought_forward", 0)])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"budget-{month}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# =================== SEED ===================

@api_router.post("/seed")
async def seed():
    """Seed initial (owner_id=None) unclaimed data. Runs once."""
    existing = await db.categories.count_documents({})
    if existing > 0:
        return {"ok": True, "seeded": False}
    month = "2026-07"
    expense_cats = [
        ("Growth", 2000), ("Food and Drinks", 1000), ("Home Internet", 78.75),
        ("Offering", 100), ("Rent", 2300), ("Transport", 100),
        ("Emergency funds", 614.93), ("Family and Friends Gift", 200),
        ("Daycare and OSC", 652.5), ("Communication", 113.35),
        ("Intertainment (Netflix)", 0), ("Fuel", 100), ("House supplies", 150),
        ("Miscellaneous", 100), ("House Furniture", 150), ("School Bus", 0),
        ("Health", 50), ("Outings", 200), ("PJ Allowance", 300),
        ("Utility", 370), ("Amazon Prime", 0), ("Car insurance", 293.1),
        ("Children supplies", 100), ("EJ Allowance", 200),
        ("Traffic Fine", 0), ("Canadian Tire", 35.83),
    ]
    income_cats = [
        ("SOFRECO", 3947.671), ("EI", 1380), ("CCB", 1580.8), ("GST/HST", 0),
        ("CCR", 0), ("ACFB", 0), ("TP", 2500), ("Debt repayment", 0),
        ("Others", 0), ("Brought forward", 0), ("Emergency Fund", 0),
    ]
    for name, planned in expense_cats:
        c = Category(owner_id=None, name=name, type="expense", planned=planned, month=month).dict()  # type: ignore
        c["owner_id"] = None
        await db.categories.insert_one(c)
    for name, planned in income_cats:
        c = Category(owner_id=None, name=name, type="income", planned=planned, month=month).dict()  # type: ignore
        c["owner_id"] = None
        await db.categories.insert_one(c)

    txs = [
        ("2026-06-29", 319.47, "Kasoa", "Food and Drinks", "expense"),
        ("2026-06-29", 333.84, "Costco", "Food and Drinks", "expense"),
        ("2026-06-29", 50.0, "Walmart", "Food and Drinks", "expense"),
        ("2026-07-05", 28.95, "H&W", "Food and Drinks", "expense"),
        ("2026-07-06", 103.97, "Footwear", "PJ Allowance", "expense"),
        ("2026-07-06", 100.24, "Footwear", "Children supplies", "expense"),
        ("2026-07-07", 1525.05, "TP Pay", "TP", "income"),
        ("2026-07-02", 3947.671, "Ginger Sofreco", "SOFRECO", "income"),
        ("2026-07-07", 690.0, "EI Payment", "EI", "income"),
    ]
    for date, amount, desc, cat, t in txs:
        d = Transaction(
            owner_id=None, date=date, amount=amount, description=desc,  # type: ignore
            category=cat, type=t, month=date[:7],
        ).dict()
        d["owner_id"] = None
        await db.transactions.insert_one(d)

    accounts = [
        ("EJ SB Current Account", "Cash", 299.27),
        ("WS Emergency", "Cash", 9843.86),
        ("CN Shares", "Investment", 4026.23),
        ("CN Shares Match", "Investment", 734.49),
        ("EJ Scotia Savings", "Cash", 125.68),
        ("EJ TFSA Portfolio", "Registered", 149.5),
        ("TFSA General", "Registered", 4868.42),
        ("Tax Cash TFSA", "Registered", 4564.16),
        ("PJ Tax TFSA", "Registered", 0.0),
        ("Non Registered Margin", "Investment", 778.91),
        ("EJ TFSA Finance", "Registered", 259.82),
        ("EJ RRSP", "Registered", 4172.61),
        ("EJ FHSA", "Registered", 6948.38),
        ("Crypto Wise", "Crypto", 170.9),
        ("RESP", "Registered", 5726.81),
        ("PJ SC Checking Account", "Cash", 527.0),
        ("PJ Savings SC", "Cash", 107.78),
        ("PJ WS CA", "Cash", 10911.07),
        ("Cash holding", "Cash", 175.0),
        ("Senegal Ecobank", "Other", 0.0),
    ]
    for name, grp, bal in accounts:
        d = Account(owner_id=None, name=name, group=grp, balance=bal, brought_forward=bal).dict()  # type: ignore
        d["owner_id"] = None
        await db.accounts.insert_one(d)

    allocations = [
        ("EF", 0.05), ("Grow", 0.15), ("FHSA", 0.30),
        ("RRSP", 0.10), ("RESP", 0.10), ("EJ TFSA", 0.10),
        ("PJ TFSA", 0.30), ("Non-Registered", 0.0), ("Crypto", 0.10),
    ]
    for name, pct in allocations:
        d = Allocation(owner_id=None, name=name, percent=pct).dict()  # type: ignore
        d["owner_id"] = None
        await db.allocations.insert_one(d)

    return {"ok": True, "seeded": True}


@api_router.get("/")
async def root():
    return {"message": "Budget API v2"}


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def on_startup():
    # Indexes
    await db.users.create_index("email", unique=True)
    await db.users.create_index("user_id", unique=True)
    await db.user_sessions.create_index("session_token", unique=True)
    await db.user_sessions.create_index("user_id")
    try:
        await db.user_sessions.create_index("expires_at", expireAfterSeconds=0)
    except Exception:
        pass
    await db.budgets.create_index("share_token")

    # Seed if empty
    try:
        count = await db.categories.count_documents({})
        if count == 0:
            await seed()
            logger.info("Auto-seeded initial data")
    except Exception as e:
        logger.error(f"Seed error: {e}")


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
